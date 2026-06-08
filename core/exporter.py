"""Генератор итогового Excel со сводной сметой и аналитикой.

Структура книги:
    1. «Сводная смета» — два блока:
        блок РАБОТЫ (по разделам клиента) и блок МАТЕРИАЛЫ (по разделам).
        Все суммы и маржа — формулами Excel. Столбцы с НДС скрыты.
    2. «Аналитика» — все цифры формулами со ссылками на лист «Сводная смета».
    3. «Исходник клиента» / «Исходник подрядчика» — оригинальные сметы
        с итогами по разделам и общим итогом, тоже формулами.

НДС РФ 2026 = 22%. Маржа считается без НДС.
"""
from __future__ import annotations

import io
import re
from datetime import date
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .models import Estimate, EstimateItem, MaterialMatch, Match


VAT_RATE = 0.22
VAT_DIVISOR = f"{1 + VAT_RATE:.2f}"

GREEN = PatternFill("solid", fgColor="C6EFCE")
YELLOW = PatternFill("solid", fgColor="FFEB9C")
RED = PatternFill("solid", fgColor="FFC7CE")
GREY = PatternFill("solid", fgColor="EEECEC")
SECTION_FILL = PatternFill("solid", fgColor="DDEBF7")
BLOCK_FILL = PatternFill("solid", fgColor="4472C4")
TOTAL_FILL = PatternFill("solid", fgColor="B4C7E7")
PROJECT_TOTAL_FILL = PatternFill("solid", fgColor="2E75B6")

BORDER_THIN = Border(
    left=Side(style="thin", color="C0C0C0"),
    right=Side(style="thin", color="C0C0C0"),
    top=Side(style="thin", color="C0C0C0"),
    bottom=Side(style="thin", color="C0C0C0"),
)
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="305496")
WHITE_BOLD = Font(bold=True, color="FFFFFF")


# ---------------------------------------------------------- Классификация материалов
CATEGORY_RULES = [
    ("Лотки", [r"лоток", r"лотка", r"разделитель лотк", r"крышка лотк",
               r"угол лотк", r"заглушк.*лотк", r"кабельнес"]),
    ("Кабель и проводка", [r"кабель", r"провод", r"витая пара",
                            r"коннектор", r"наконечник", r"муфт"]),
    ("Зарядные станции", [r"эзс", r"зарядн", r"пур[\-\s]?эзс", r"щсу[\-\s]?эзс"]),
    ("Шкафы и щиты", [r"шкаф", r"щит", r"сдупэм", r"щсу", r"корпус"]),
]


def classify_material(name: str) -> str:
    n = name.lower()
    for category, patterns in CATEGORY_RULES:
        for pat in patterns:
            if re.search(pat, n):
                return category
    return "Другое"


def _margin_fill(margin_pct: Optional[float]) -> PatternFill:
    if margin_pct is None:
        return GREY
    if margin_pct >= 0.25:
        return GREEN
    if margin_pct >= 0.10:
        return YELLOW
    return RED


# ---------------------------------------------------------- Колонки листа «Сводная смета»
# Порядок: №, Наименование клиента, Ед., Кол-во,
#          Цена/ед. с НДС (скрыт), Цена/ед. без НДС,
#          Сумма с НДС (скрыт), Сумма без НДС,
#          Цена закупки (без НДС), Сумма закупки (без НДС),
#          Наименование у подрядчика, Исполнитель / Поставщик,
#          Маржа руб, Маржа %
COLUMNS = [
    ("№", 6),
    ("Наименование (верхняя смета)", 50),
    ("Ед.", 8),
    ("Кол-во", 10),
    ("Цена/ед. (с НДС)", 14),
    ("Цена/ед. (без НДС)", 14),
    ("Сумма (с НДС)", 16),
    ("Сумма (без НДС)", 16),
    ("Цена закупки (без НДС)", 16),
    ("Сумма закупки (без НДС)", 18),
    ("Наименование (нижняя смета)", 50),
    ("Исполнитель / Поставщик", 24),
    ("Маржа, руб.", 14),
    ("Маржа, %", 11),
]
N = len(COLUMNS)

# Удобные индексы
NUMBER = 1
NAME_CLIENT = 2
UNIT = 3
QTY = 4
UNIT_PRICE_GROSS = 5      # скрыт
UNIT_PRICE_NET = 6
SUM_GROSS = 7              # скрыт
SUM_NET = 8
UNIT_PRICE_PURCHASE = 9
SUM_PURCHASE = 10
NAME_BOTTOM = 11
EXECUTOR = 12
MARGIN_ABS = 13
MARGIN_PCT = 14

HIDDEN_COLUMNS = (UNIT_PRICE_GROSS, SUM_GROSS)


def _L(col: int) -> str:
    return get_column_letter(col)


def _money_format(cell):
    cell.number_format = "#,##0.00 ₽"


def _pct_format(cell):
    cell.number_format = "0.0%"


def _border_row(ws, row: int, ncols: int):
    for c in range(1, ncols + 1):
        ws.cell(row=row, column=c).border = BORDER_THIN


def _write_headers(ws: Worksheet):
    for col_idx, (title, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER_THIN
        ws.column_dimensions[_L(col_idx)].width = width
    for c in HIDDEN_COLUMNS:
        ws.column_dimensions[_L(c)].hidden = True
    ws.row_dimensions[1].height = 40
    ws.freeze_panes = f"{_L(NAME_CLIENT + 1)}2"


def _block_header_row(ws, row: int, title: str):
    for c in range(1, N + 1):
        ws.cell(row=row, column=c).fill = BLOCK_FILL
        ws.cell(row=row, column=c).font = WHITE_BOLD
        ws.cell(row=row, column=c).border = BORDER_THIN
    ws.cell(row=row, column=1, value=title).alignment = Alignment(horizontal="left",
                                                                     vertical="center")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=N)
    ws.row_dimensions[row].height = 22


def _section_header_row(ws, row: int, section: str):
    for c in range(1, N + 1):
        ws.cell(row=row, column=c).fill = SECTION_FILL
        ws.cell(row=row, column=c).border = BORDER_THIN
    ws.cell(row=row, column=1, value=section).font = Font(bold=True)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=N)


def _write_item_row(
    ws,
    row: int,
    item: EstimateItem,
    kind: str,                            # "work" или "material"
    contractor_item: Optional[EstimateItem],
    contractor_title: str,
    *,
    supplier_row: Optional[dict] = None,  # позиция из счёта поставщика (для материалов)
):
    """Записывает строку. kind определяет какую цену клиента взять (мат/раб).

    supplier_row — позиция счёта поставщика (dict из db.all_invoice_items_for_project)
    с полями: name, article_supplier, article_manufacturer, unit_price, vat_included,
    vat_rate, supplier_name, invoice_number, invoice_date. Если задан и kind=="material",
    эти данные пойдут в колонки «закупка» вместо подрядчика."""
    qty = item.quantity or 0
    unit_price_gross = (item.price_work if kind == "work" else item.price_material)

    ws.cell(row=row, column=NUMBER, value=item.number)
    ws.cell(row=row, column=NAME_CLIENT, value=item.name).alignment = Alignment(
        wrap_text=True, vertical="top"
    )
    ws.cell(row=row, column=UNIT, value=item.unit)
    ws.cell(row=row, column=QTY, value=qty).number_format = "#,##0.00"

    # Цена с НДС (скрытая колонка) — значение
    if unit_price_gross is not None:
        ws.cell(row=row, column=UNIT_PRICE_GROSS, value=unit_price_gross)
        _money_format(ws.cell(row=row, column=UNIT_PRICE_GROSS))

        # Цена без НДС — формула
        ws.cell(row=row, column=UNIT_PRICE_NET,
                value=f"={_L(UNIT_PRICE_GROSS)}{row}/{VAT_DIVISOR}")
        _money_format(ws.cell(row=row, column=UNIT_PRICE_NET))

        # Сумма с НДС (скрытая) — qty * unit_gross
        ws.cell(row=row, column=SUM_GROSS,
                value=f"={_L(QTY)}{row}*{_L(UNIT_PRICE_GROSS)}{row}")
        _money_format(ws.cell(row=row, column=SUM_GROSS))

        # Сумма без НДС — формула qty * unit_net
        ws.cell(row=row, column=SUM_NET,
                value=f"={_L(QTY)}{row}*{_L(UNIT_PRICE_NET)}{row}")
        _money_format(ws.cell(row=row, column=SUM_NET))

    # ----- Закупка -----
    purchase_unit_price_net: Optional[float] = None
    bottom_name: Optional[str] = None
    executor_label: Optional[str] = None

    if kind == "work" and contractor_item is not None \
            and contractor_item.price_work is not None:
        # Цены подрядчика в нашем парсере уже без НДС
        purchase_unit_price_net = contractor_item.price_work
        bottom_name = contractor_item.name
        executor_label = contractor_title
    elif kind == "material" and supplier_row is not None \
            and supplier_row.get("unit_price") is not None:
        # Цена из счёта поставщика. В счёте может быть с НДС или без.
        raw_price = float(supplier_row["unit_price"])
        if supplier_row.get("vat_included"):
            vat_rate = supplier_row.get("vat_rate") or VAT_RATE
            purchase_unit_price_net = raw_price / (1 + vat_rate)
        else:
            purchase_unit_price_net = raw_price

        # Имя в нижней строке — позиция из счёта + артикул (если есть)
        art = (supplier_row.get("article_supplier")
                or supplier_row.get("article_manufacturer"))
        name_with_art = supplier_row["name"]
        if art:
            name_with_art = f"[{art}] {name_with_art}"
        bottom_name = name_with_art

        sup_name = supplier_row.get("supplier_name") or "—"
        inv_no = supplier_row.get("invoice_number") or "—"
        inv_dt = supplier_row.get("invoice_date") or ""
        executor_label = f"{sup_name} • счёт №{inv_no}"
        if inv_dt:
            executor_label += f" от {inv_dt}"

    if purchase_unit_price_net is not None:
        ws.cell(row=row, column=UNIT_PRICE_PURCHASE, value=purchase_unit_price_net)
        _money_format(ws.cell(row=row, column=UNIT_PRICE_PURCHASE))

        ws.cell(row=row, column=SUM_PURCHASE,
                value=f"={_L(QTY)}{row}*{_L(UNIT_PRICE_PURCHASE)}{row}")
        _money_format(ws.cell(row=row, column=SUM_PURCHASE))

        ws.cell(row=row, column=NAME_BOTTOM, value=bottom_name).alignment = \
            Alignment(wrap_text=True, vertical="top")
        ws.cell(row=row, column=EXECUTOR, value=executor_label)

        # Маржа
        ws.cell(row=row, column=MARGIN_ABS,
                value=f"={_L(SUM_NET)}{row}-{_L(SUM_PURCHASE)}{row}")
        _money_format(ws.cell(row=row, column=MARGIN_ABS))

        pct_cell = ws.cell(row=row, column=MARGIN_PCT,
                            value=f"=IF({_L(SUM_NET)}{row}=0,0,"
                                  f"{_L(MARGIN_ABS)}{row}/{_L(SUM_NET)}{row})")
        _pct_format(pct_cell)

        # Светофор маржи — статически по нашим значениям
        if unit_price_gross and qty:
            client_net = qty * unit_price_gross / (1 + VAT_RATE)
            purchase_net = qty * purchase_unit_price_net
            mp = (client_net - purchase_net) / client_net if client_net else None
            pct_cell.fill = _margin_fill(mp)
    else:
        # помечаем пустые колонки маржи
        if kind == "material":
            ws.cell(row=row, column=NAME_BOTTOM, value="— нет цены закупки —")
            ws.cell(row=row, column=EXECUTOR, value="(материал)").fill = GREY
        else:
            ws.cell(row=row, column=NAME_BOTTOM, value="— не закрыто —")
            ws.cell(row=row, column=EXECUTOR, value="—").fill = GREY
        ws.cell(row=row, column=MARGIN_PCT).fill = GREY

    _border_row(ws, row, N)


def _section_subtotal_row(ws, row: int, label: str,
                           item_rows: list[int]) -> dict:
    """Итог по разделу — формулами SUM. Возвращает адреса для аналитики."""
    if not item_rows:
        return {}
    first, last = item_rows[0], item_rows[-1]
    for c in range(1, N + 1):
        ws.cell(row=row, column=c).fill = TOTAL_FILL
        ws.cell(row=row, column=c).font = Font(bold=True)
        ws.cell(row=row, column=c).border = BORDER_THIN
    ws.cell(row=row, column=1, value=label)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=QTY)

    ws.cell(row=row, column=SUM_GROSS,
            value=f"=SUM({_L(SUM_GROSS)}{first}:{_L(SUM_GROSS)}{last})")
    _money_format(ws.cell(row=row, column=SUM_GROSS))

    ws.cell(row=row, column=SUM_NET,
            value=f"=SUM({_L(SUM_NET)}{first}:{_L(SUM_NET)}{last})")
    _money_format(ws.cell(row=row, column=SUM_NET))

    ws.cell(row=row, column=SUM_PURCHASE,
            value=f"=SUM({_L(SUM_PURCHASE)}{first}:{_L(SUM_PURCHASE)}{last})")
    _money_format(ws.cell(row=row, column=SUM_PURCHASE))

    ws.cell(row=row, column=MARGIN_ABS,
            value=f"={_L(SUM_NET)}{row}-{_L(SUM_PURCHASE)}{row}")
    _money_format(ws.cell(row=row, column=MARGIN_ABS))

    ws.cell(row=row, column=MARGIN_PCT,
            value=f"=IF({_L(SUM_NET)}{row}=0,0,"
                  f"{_L(MARGIN_ABS)}{row}/{_L(SUM_NET)}{row})")
    _pct_format(ws.cell(row=row, column=MARGIN_PCT))

    return {
        "row": row,
        "sum_gross": f"{_L(SUM_GROSS)}{row}",
        "sum_net": f"{_L(SUM_NET)}{row}",
        "sum_purchase": f"{_L(SUM_PURCHASE)}{row}",
    }


def _block_total_row(ws, row: int, label: str,
                      section_subtotals: list[dict], fill=PROJECT_TOTAL_FILL,
                      font=WHITE_BOLD) -> dict:
    """Итог по блоку (Работы / Материалы / Проект)."""
    for c in range(1, N + 1):
        ws.cell(row=row, column=c).fill = fill
        ws.cell(row=row, column=c).font = font
        ws.cell(row=row, column=c).border = BORDER_THIN
    ws.cell(row=row, column=1, value=label)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=QTY)

    if not section_subtotals:
        return {"row": row}

    for col_key, target_col in [
        ("sum_gross", SUM_GROSS),
        ("sum_net", SUM_NET),
        ("sum_purchase", SUM_PURCHASE),
    ]:
        formula = "+".join(s[col_key] for s in section_subtotals)
        cell = ws.cell(row=row, column=target_col, value=f"={formula}")
        _money_format(cell)

    ws.cell(row=row, column=MARGIN_ABS,
            value=f"={_L(SUM_NET)}{row}-{_L(SUM_PURCHASE)}{row}")
    _money_format(ws.cell(row=row, column=MARGIN_ABS))

    ws.cell(row=row, column=MARGIN_PCT,
            value=f"=IF({_L(SUM_NET)}{row}=0,0,"
                  f"{_L(MARGIN_ABS)}{row}/{_L(SUM_NET)}{row})")
    _pct_format(ws.cell(row=row, column=MARGIN_PCT))

    return {
        "row": row,
        "sum_gross": f"{_L(SUM_GROSS)}{row}",
        "sum_net": f"{_L(SUM_NET)}{row}",
        "sum_purchase": f"{_L(SUM_PURCHASE)}{row}",
    }


def _build_summary_sheet(ws: Worksheet, client: Estimate, contractor: Estimate,
                          matches: list[Match],
                          material_matches: Optional[list[MaterialMatch]] = None,
                          supplier_rows: Optional[list[dict]] = None):
    ws.title = "Сводная смета"
    _write_headers(ws)

    match_by_client = {m.client_idx: m for m in matches}

    # Карта: индекс клиентской позиции -> dict позиции счёта (если найден матч)
    supplier_by_client_idx: dict[int, dict] = {}
    if material_matches and supplier_rows:
        supplier_by_id = {r["id"]: r for r in supplier_rows}
        for mm in material_matches:
            if mm.invoice_item_id is not None and mm.confidence >= 0.4:
                row = supplier_by_id.get(mm.invoice_item_id)
                if row:
                    supplier_by_client_idx[mm.client_idx] = row

    sections_order: list[str] = []
    seen = set()
    for it in client.items:
        if it.section not in seen:
            sections_order.append(it.section)
            seen.add(it.section)

    # Подготовка списков работ и материалов по разделам
    work_items_by_section: dict[str, list[tuple[int, EstimateItem]]] = {}
    material_items_by_section: dict[str, list[tuple[int, EstimateItem]]] = {}

    for c_idx, it in enumerate(client.items):
        has_work = it.price_work is not None or it.sum_work is not None
        has_mat = it.price_material is not None or it.sum_material is not None
        if has_work:
            work_items_by_section.setdefault(it.section, []).append((c_idx, it))
        if has_mat:
            material_items_by_section.setdefault(it.section, []).append((c_idx, it))

    state = {
        "work_section_subtotals": {},   # section -> dict (row, sum_net, sum_purchase ...)
        "material_section_subtotals": {},
        "work_block_total": None,
        "material_block_total": None,
        "project_total": None,
    }

    row = 2

    # ============================== БЛОК РАБОТЫ
    _block_header_row(ws, row, "🔧 РАБОТЫ")
    row += 1

    work_section_addrs: list[dict] = []
    for section in sections_order:
        items = work_items_by_section.get(section, [])
        if not items:
            continue
        _section_header_row(ws, row, section)
        row += 1
        item_rows: list[int] = []
        for c_idx, it in items:
            match = match_by_client.get(c_idx)
            contractor_item = (
                contractor.items[match.contractor_idx]
                if match and match.contractor_idx is not None else None
            )
            _write_item_row(ws, row, it, "work", contractor_item, contractor.title)
            item_rows.append(row)
            row += 1
        subtotal = _section_subtotal_row(
            ws, row, f"Итого работ по разделу: {section}", item_rows,
        )
        if subtotal:
            state["work_section_subtotals"][section] = subtotal
            work_section_addrs.append(subtotal)
        row += 1

    state["work_block_total"] = _block_total_row(
        ws, row, "ИТОГО ПО РАБОТАМ", work_section_addrs,
        fill=PROJECT_TOTAL_FILL, font=WHITE_BOLD,
    )
    row += 2

    # ============================== БЛОК МАТЕРИАЛЫ
    _block_header_row(ws, row, "📦 МАТЕРИАЛЫ")
    row += 1

    material_section_addrs: list[dict] = []
    for section in sections_order:
        items = material_items_by_section.get(section, [])
        if not items:
            continue
        _section_header_row(ws, row, section)
        row += 1
        item_rows: list[int] = []
        for c_idx, it in items:
            sup_row = supplier_by_client_idx.get(c_idx)
            _write_item_row(
                ws, row, it, "material", None, contractor.title,
                supplier_row=sup_row,
            )
            item_rows.append(row)
            row += 1
        subtotal = _section_subtotal_row(
            ws, row, f"Итого материалов по разделу: {section}", item_rows,
        )
        if subtotal:
            state["material_section_subtotals"][section] = subtotal
            material_section_addrs.append(subtotal)
        row += 1

    state["material_block_total"] = _block_total_row(
        ws, row, "ИТОГО ПО МАТЕРИАЛАМ", material_section_addrs,
        fill=PROJECT_TOTAL_FILL, font=WHITE_BOLD,
    )
    row += 2

    # ============================== ВСЕГО ПО ПРОЕКТУ
    project_parts = []
    if state["work_block_total"] and "sum_net" in state["work_block_total"]:
        project_parts.append(state["work_block_total"])
    if state["material_block_total"] and "sum_net" in state["material_block_total"]:
        project_parts.append(state["material_block_total"])

    state["project_total"] = _block_total_row(
        ws, row, "ВСЕГО ПО ПРОЕКТУ", project_parts,
        fill=PROJECT_TOTAL_FILL, font=WHITE_BOLD,
    )
    ws.row_dimensions[row].height = 22

    return state


# ---------------------------------------------------------- Аналитика
SUMMARY_SHEET = "'Сводная смета'"


def _build_analytics_sheet(ws: Worksheet, client: Estimate, contractor: Estimate,
                            matches: list[Match], summary_state: dict):
    ws.title = "Аналитика"
    for col, w in enumerate([42, 22, 22, 18, 14, 26], start=1):
        ws.column_dimensions[_L(col)].width = w
    ws.freeze_panes = "A2"

    work_subs = summary_state.get("work_section_subtotals", {})
    mat_subs = summary_state.get("material_section_subtotals", {})
    work_total = summary_state.get("work_block_total")
    mat_total = summary_state.get("material_block_total")
    project_total = summary_state.get("project_total")

    row = 1
    ws.cell(row=row, column=1,
            value="📊 Аналитика проекта (все цифры — формулы со ссылками на «Сводную смету»)").font = \
        Font(bold=True, size=12)
    row += 2

    # ----- метрики проекта
    ws.cell(row=row, column=1, value="Показатель").font = HEADER_FONT
    ws.cell(row=row, column=2, value="Значение").font = HEADER_FONT
    for c in (1, 2):
        ws.cell(row=row, column=c).fill = HEADER_FILL
        ws.cell(row=row, column=c).border = BORDER_THIN
    row += 1

    def _ref(addr_dict, key, default="0"):
        if addr_dict and key in addr_dict:
            return f"={SUMMARY_SHEET}!{addr_dict[key]}"
        return f"={default}"

    metrics = [
        ("Сумма проекта у клиента (с НДС 22%)",
         _ref(project_total, "sum_gross"), "#,##0.00 ₽", None),
        ("Сумма проекта у клиента (без НДС)",
         _ref(project_total, "sum_net"), "#,##0.00 ₽", None),
        ("Сумма закупки работ у подрядчика (без НДС)",
         _ref(work_total, "sum_purchase"), "#,##0.00 ₽", None),
        ("Маржа по работам (руб., без НДС)",
         f"={SUMMARY_SHEET}!{work_total['sum_net']}-{SUMMARY_SHEET}!{work_total['sum_purchase']}"
         if work_total else "0", "#,##0.00 ₽", None),
        ("Маржа по работам (%)",
         f"=IF({SUMMARY_SHEET}!{work_total['sum_net']}=0,0,"
         f"({SUMMARY_SHEET}!{work_total['sum_net']}-{SUMMARY_SHEET}!{work_total['sum_purchase']})"
         f"/{SUMMARY_SHEET}!{work_total['sum_net']})"
         if work_total else "0", "0.0%", "margin"),
        ("Позиций у клиента", len(client.items), "0", None),
        ("Из них с матчем подрядчика",
         sum(1 for m in matches if m.contractor_idx is not None), "0", None),
    ]
    for label, value, fmt, hue in metrics:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        cell = ws.cell(row=row, column=2, value=value)
        cell.number_format = fmt
        # Светофор маржи по работам — оценим по нашим Python-данным
        if hue == "margin":
            work_client = sum(
                ((it.sum_work or 0) / (1 + VAT_RATE))
                for it in client.items
                if it.sum_work is not None
            )
            work_purchase = 0.0
            for c_idx, it in enumerate(client.items):
                if it.kind not in ("work", "mixed", "composite"):
                    continue
                m = next((mm for mm in matches if mm.client_idx == c_idx), None)
                if m and m.contractor_idx is not None:
                    cp = contractor.items[m.contractor_idx]
                    work_purchase += (cp.price_work or 0) * (it.quantity or 0)
            mp = ((work_client - work_purchase) / work_client) if work_client else None
            cell.fill = _margin_fill(mp)
        for c in (1, 2):
            ws.cell(row=row, column=c).border = BORDER_THIN
        row += 1
    row += 2

    # ----- Маржа по работам — по разделам
    ws.cell(row=row, column=1,
            value="🔧 Маржа по работам по разделам").font = Font(bold=True, size=12)
    row += 1
    headers = ["Раздел", "Клиент (без НДС)", "Подрядчик (без НДС)",
               "Маржа, руб.", "Маржа, %", "Статус"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.border = BORDER_THIN
    row += 1

    work_first_row = row
    sections_order: list[str] = []
    seen = set()
    for it in client.items:
        if it.section not in seen:
            sections_order.append(it.section)
            seen.add(it.section)

    # счётчик матчей по разделам — для определения статуса
    matched_counts: dict[str, dict] = {s: {"items": 0, "matched": 0} for s in sections_order}
    for c_idx, it in enumerate(client.items):
        if it.kind not in ("work", "mixed", "composite"):
            continue
        if it.sum_work is None and it.price_work is None:
            continue
        matched_counts[it.section]["items"] += 1
        m = next((mm for mm in matches if mm.client_idx == c_idx), None)
        if m and m.contractor_idx is not None:
            matched_counts[it.section]["matched"] += 1

    for section in sections_order:
        sub = work_subs.get(section)
        if sub is None:
            continue
        ws.cell(row=row, column=1, value=section)
        ws.cell(row=row, column=2,
                value=f"={SUMMARY_SHEET}!{sub['sum_net']}").number_format = "#,##0.00 ₽"
        ws.cell(row=row, column=3,
                value=f"={SUMMARY_SHEET}!{sub['sum_purchase']}").number_format = "#,##0.00 ₽"
        ws.cell(row=row, column=4,
                value=f"=B{row}-C{row}").number_format = "#,##0.00 ₽"
        pct = ws.cell(row=row, column=5,
                       value=f"=IF(B{row}=0,0,D{row}/B{row})")
        pct.number_format = "0.0%"

        cnt = matched_counts.get(section, {"items": 0, "matched": 0})
        if cnt["matched"] == 0:
            status = "⚠ Нет нижних цен"
            pct.fill = GREY
        elif cnt["matched"] < cnt["items"]:
            status = f"Частично: {cnt['matched']} из {cnt['items']}"
        else:
            status = f"Закрыто ✓ ({cnt['items']} поз.)"
        ws.cell(row=row, column=6, value=status)
        for c in range(1, 7):
            ws.cell(row=row, column=c).border = BORDER_THIN
        row += 1

    # Итого по работам
    if work_first_row < row and work_total:
        for c in range(1, 7):
            ws.cell(row=row, column=c).fill = TOTAL_FILL
            ws.cell(row=row, column=c).font = Font(bold=True)
            ws.cell(row=row, column=c).border = BORDER_THIN
        ws.cell(row=row, column=1, value="Итого по работам")
        ws.cell(row=row, column=2,
                value=f"={SUMMARY_SHEET}!{work_total['sum_net']}").number_format = "#,##0.00 ₽"
        ws.cell(row=row, column=3,
                value=f"={SUMMARY_SHEET}!{work_total['sum_purchase']}").number_format = "#,##0.00 ₽"
        ws.cell(row=row, column=4, value=f"=B{row}-C{row}").number_format = "#,##0.00 ₽"
        ws.cell(row=row, column=5,
                value=f"=IF(B{row}=0,0,D{row}/B{row})").number_format = "0.0%"
        row += 1
    row += 2

    # ----- Маржа по материалам — по категориям
    # Категоризируем материалы из клиентской сметы и связываем с строками на «Сводной смете»
    # Для аналитики берём суммы из «Сводной смете» — но категории нужно агрегировать.
    # Сначала строим адреса строк материалов на листе «Сводная смета».
    # У нас на «Сводной смете» материалы лежат в блоке МАТЕРИАЛЫ по разделам.
    # Для категорий — пересчитаем по нашим данным (без НДС) и положим как формулы /1.22 в Excel.

    ws.cell(row=row, column=1,
            value="📦 Материалы по категориям").font = Font(bold=True, size=12)
    row += 1
    headers = ["Категория", "Клиент (без НДС)", "Закупка (без НДС)",
               "Маржа, руб.", "Маржа, %", "Статус"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.border = BORDER_THIN
    row += 1

    # агрегаты по категориям (без НДС, на клиенте)
    cat_totals: dict[str, float] = {}
    cat_counts: dict[str, int] = {}
    for it in client.items:
        if it.price_material is None and it.sum_material is None:
            continue
        sum_mat_gross = it.sum_material if it.sum_material is not None \
            else (it.price_material or 0) * (it.quantity or 0)
        sum_mat_net = sum_mat_gross / (1 + VAT_RATE)
        category = classify_material(it.name)
        cat_totals[category] = cat_totals.get(category, 0.0) + sum_mat_net
        cat_counts[category] = cat_counts.get(category, 0) + 1

    mat_first_row = row
    for category in ["Лотки", "Кабель и проводка", "Зарядные станции",
                     "Шкафы и щиты", "Другое"]:
        if category not in cat_totals:
            continue
        ws.cell(row=row, column=1, value=category)
        ws.cell(row=row, column=2,
                value=cat_totals[category]).number_format = "#,##0.00 ₽"
        ws.cell(row=row, column=3, value=0).number_format = "#,##0.00 ₽"
        ws.cell(row=row, column=4,
                value=f"=B{row}-C{row}").number_format = "#,##0.00 ₽"
        pct = ws.cell(row=row, column=5,
                       value=f"=IF(B{row}=0,0,D{row}/B{row})")
        pct.number_format = "0.0%"
        pct.fill = GREY
        ws.cell(row=row, column=6,
                value=f"⚠ Нет цен закупки ({cat_counts[category]} поз.)").font = \
            Font(italic=True)
        for c in range(1, 7):
            ws.cell(row=row, column=c).border = BORDER_THIN
        row += 1

    if mat_first_row < row:
        last = row - 1
        for c in range(1, 7):
            ws.cell(row=row, column=c).fill = TOTAL_FILL
            ws.cell(row=row, column=c).font = Font(bold=True)
            ws.cell(row=row, column=c).border = BORDER_THIN
        ws.cell(row=row, column=1, value="Итого по материалам")
        ws.cell(row=row, column=2,
                value=f"=SUM(B{mat_first_row}:B{last})").number_format = "#,##0.00 ₽"
        ws.cell(row=row, column=3,
                value=f"=SUM(C{mat_first_row}:C{last})").number_format = "#,##0.00 ₽"
        ws.cell(row=row, column=4,
                value=f"=B{row}-C{row}").number_format = "#,##0.00 ₽"
        ws.cell(row=row, column=5,
                value=f"=IF(B{row}=0,0,D{row}/B{row})").number_format = "0.0%"


# ---------------------------------------------------------- Исходники с итогами
def _build_raw_sheet(ws: Worksheet, estimate: Estimate, title: str):
    ws.title = title[:31]
    headers = ["Раздел", "№", "Тип", "Наименование", "Ед.", "Кол-во",
               "Цена мат.", "Цена раб.", "Сумма мат.", "Сумма раб.", "Итого"]
    widths = [28, 5, 12, 55, 8, 10, 14, 14, 18, 18, 18]
    ncols = len(headers)
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.border = BORDER_THIN
        ws.column_dimensions[_L(i)].width = w
    ws.freeze_panes = "A2"

    sections_order: list[str] = []
    seen = set()
    for it in estimate.items:
        if it.section not in seen:
            sections_order.append(it.section)
            seen.add(it.section)

    row = 2
    section_total_rows: list[int] = []

    for section in sections_order:
        for c in range(1, ncols + 1):
            ws.cell(row=row, column=c).fill = SECTION_FILL
            ws.cell(row=row, column=c).border = BORDER_THIN
        ws.cell(row=row, column=1, value=section).font = Font(bold=True)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
        row += 1

        item_rows = []
        for it in estimate.items:
            if it.section != section:
                continue
            ws.cell(row=row, column=1, value=it.section)
            ws.cell(row=row, column=2, value=it.number)
            ws.cell(row=row, column=3, value=it.kind)
            ws.cell(row=row, column=4, value=it.name).alignment = \
                Alignment(wrap_text=True, vertical="top")
            ws.cell(row=row, column=5, value=it.unit)
            ws.cell(row=row, column=6, value=it.quantity).number_format = "#,##0.00"

            qty_ref = f"F{row}"
            if it.price_material is not None:
                ws.cell(row=row, column=7, value=it.price_material).number_format = "#,##0.00"
            if it.price_work is not None:
                ws.cell(row=row, column=8, value=it.price_work).number_format = "#,##0.00"
            if it.price_material is not None and it.quantity is not None:
                ws.cell(row=row, column=9, value=f"={qty_ref}*G{row}").number_format = "#,##0.00"
            elif it.sum_material is not None:
                ws.cell(row=row, column=9, value=it.sum_material).number_format = "#,##0.00"
            if it.price_work is not None and it.quantity is not None:
                ws.cell(row=row, column=10, value=f"={qty_ref}*H{row}").number_format = "#,##0.00"
            elif it.sum_work is not None:
                ws.cell(row=row, column=10, value=it.sum_work).number_format = "#,##0.00"
            ws.cell(row=row, column=11,
                    value=f"=IFERROR(I{row},0)+IFERROR(J{row},0)").number_format = "#,##0.00"

            for c in range(1, ncols + 1):
                ws.cell(row=row, column=c).border = BORDER_THIN
            item_rows.append(row)
            row += 1

        if item_rows:
            first = item_rows[0]
            last = item_rows[-1]
            for c in range(1, ncols + 1):
                ws.cell(row=row, column=c).fill = TOTAL_FILL
                ws.cell(row=row, column=c).font = Font(bold=True)
                ws.cell(row=row, column=c).border = BORDER_THIN
            ws.cell(row=row, column=1, value=f"Итого: {section}")
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
            ws.cell(row=row, column=9,
                    value=f"=SUM(I{first}:I{last})").number_format = "#,##0.00"
            ws.cell(row=row, column=10,
                    value=f"=SUM(J{first}:J{last})").number_format = "#,##0.00"
            ws.cell(row=row, column=11,
                    value=f"=SUM(K{first}:K{last})").number_format = "#,##0.00"
            section_total_rows.append(row)
            row += 1
        row += 1

    if section_total_rows:
        row += 1
        for c in range(1, ncols + 1):
            ws.cell(row=row, column=c).fill = PROJECT_TOTAL_FILL
            ws.cell(row=row, column=c).font = WHITE_BOLD
            ws.cell(row=row, column=c).border = BORDER_THIN
        ws.cell(row=row, column=1, value=f"ИТОГО ПО СМЕТЕ: {estimate.title}")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        for col, letter in [(9, "I"), (10, "J"), (11, "K")]:
            parts = "+".join(f"{letter}{tr}" for tr in section_total_rows)
            ws.cell(row=row, column=col, value=f"={parts}").number_format = "#,##0.00"


# ---------------------------------------------------------- Сборка
def build_workbook(client: Estimate, contractor: Estimate,
                    matches: list[Match],
                    material_matches: Optional[list[MaterialMatch]] = None,
                    supplier_rows: Optional[list[dict]] = None) -> io.BytesIO:
    """Собирает Excel со сводной сметой.

    material_matches + supplier_rows — опционально. Если переданы, в строках
    материалов будут проставлены цена закупки, поставщик и № счёта, а маржа
    по материалам посчитается формулами.
    """
    wb = Workbook()
    summary_state = _build_summary_sheet(
        wb.active, client, contractor, matches,
        material_matches=material_matches,
        supplier_rows=supplier_rows,
    )
    _build_analytics_sheet(wb.create_sheet(), client, contractor, matches, summary_state)
    _build_raw_sheet(wb.create_sheet(), client, "Исходник клиента")
    _build_raw_sheet(wb.create_sheet(), contractor, "Исходник подрядчика")
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def filename_for(client: Estimate) -> str:
    safe = "".join(c if c.isalnum() or c in " -_." else "_" for c in client.title)
    return f"Сводная смета — {safe} — {date.today().isoformat()}.xlsx"
