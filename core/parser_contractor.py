"""Парсер сметы подрядчика (формат типа OPUS Okin).

Структура:
    - 3-5 строк шапки (приложение, договор, заголовок)
    - Строка заголовков: №, Наименование, Ед. изм., Кол-во, Стоимость, цена
    - Разделы: строка только с названием во 2-й колонке (нет номера, нет цен)
    - Позиции: с номером в 1-й колонке
    - Итог по разделу: "ИТОГО:" во 2-й колонке
    - Финал: "Всего по спецификации"
    - НДС не включён (договариваемся по умолчанию).
"""
from __future__ import annotations

from pathlib import Path

import openpyxl

from .models import Estimate, EstimateItem


HEADER_KEYWORDS = ("№", "Наименование")
SECTION_TERMINATORS = ("итого", "всего", "подпис", "от ", "м.п.")


def _looks_like_header(row: tuple) -> bool:
    cells = [str(c).strip().lower() if c is not None else "" for c in row]
    return "№" in cells and any("наименован" in c for c in cells)


def _is_total_row(name: str) -> bool:
    n = name.strip().lower()
    return n.startswith("итого") or n.startswith("всего")


def _to_float(value) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(str(value).replace(",", ".").replace(" ", ""))
    except (ValueError, TypeError):
        return None


def parse_contractor_estimate(file_path: str | Path) -> Estimate:
    path = Path(file_path)
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    estimate = Estimate(
        source="contractor",
        title=path.stem,
        file_name=path.name,
        vat_included=False,
    )

    rows = list(ws.iter_rows(values_only=True))

    header_idx = next(
        (i for i, r in enumerate(rows) if _looks_like_header(r)),
        None,
    )
    if header_idx is None:
        raise ValueError(
            "Не нашёл строку с заголовками таблицы (№, Наименование...)"
        )

    current_section = "Без раздела"

    for idx, row in enumerate(rows[header_idx + 1:], start=header_idx + 1):
        num, name, unit, qty, total_sum, unit_price = (row + (None,) * 6)[:6]

        if not name:
            continue

        name_str = str(name).strip()
        name_lower = name_str.lower()

        if name_lower.startswith("всего по спецификации"):
            estimate.total = _to_float(total_sum)
            continue

        if _is_total_row(name_str):
            continue

        if num is None and unit is None and qty is None:
            current_section = name_str
            continue

        item = EstimateItem(
            section=current_section,
            number=str(num) if num is not None else "",
            name=name_str,
            unit=str(unit) if unit else "",
            quantity=_to_float(qty) or 0,
            kind="work",
            price_work=_to_float(unit_price),
            sum_work=_to_float(total_sum),
            sum_total=_to_float(total_sum),
            vat_included=False,
            raw_row_index=idx + 1,
        )
        estimate.items.append(item)

    return estimate
