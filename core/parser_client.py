"""Парсер сметы клиента (формат типа OPUS FODD).

Структура:
    - Шапка (Приложение, Сводный сметный расчёт, объект)
    - Заголовки: №, Наименование, Ед., Кол-во,
        Цена материала, Цена работы, Итого ед.,
        Сумма материала, Сумма работа, ИТОГО
    - Разделы: вложенный формат "1.1 СМР в ГРЩ", "2. ПНР"
    - Позиции: с номером
    - Составные позиции (шкафы): основная строка с ценой,
        ниже строки-компоненты без номера и без цены
    - НДС включён.
"""
from __future__ import annotations

from pathlib import Path

import openpyxl

from .models import Estimate, EstimateItem


def _to_float(value) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(str(value).replace(",", ".").replace(" ", ""))
    except (ValueError, TypeError):
        return None


def _looks_like_header(row: tuple) -> bool:
    cells = [str(c).strip().lower() if c is not None else "" for c in row]
    return "№" in cells and any("наименован" in c for c in cells)


def _is_summary_row(name: str) -> bool:
    n = name.strip().lower()
    return n.startswith(("итого", "всего", "подпис", "м.п"))


def _is_section_header(row: tuple) -> bool:
    """Раздел: номер пуст, есть только название во 2-й колонке, нет цен/количества."""
    num, name = row[0], row[1]
    unit, qty = row[2], row[3]
    if not name:
        return False
    if num is not None:
        return False
    if unit is not None or qty is not None:
        return False
    return True


def parse_client_estimate(file_path: str | Path) -> Estimate:
    path = Path(file_path)
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    estimate = Estimate(
        source="client",
        title=path.stem,
        file_name=path.name,
        vat_included=True,
    )

    rows = list(ws.iter_rows(values_only=True))

    header_idx = next(
        (i for i, r in enumerate(rows) if _looks_like_header(r)),
        None,
    )
    if header_idx is None:
        raise ValueError(
            "Не нашёл строку с заголовками таблицы (№, Наименование...)"
        )

    current_section = "Без раздела"
    current_composite: EstimateItem | None = None

    for idx, row in enumerate(rows[header_idx + 1:], start=header_idx + 1):
        padded = (row + (None,) * 10)[:10]
        (
            num, name, unit, qty,
            price_mat, price_work, price_total_unit,
            sum_mat, sum_work, sum_total,
        ) = padded

        if not name:
            continue

        name_str = str(name).strip()

        if name_str.lower().startswith("итого, руб"):
            estimate.total = _to_float(sum_total)
            current_composite = None
            continue

        if _is_summary_row(name_str):
            current_composite = None
            continue

        if num is None and current_composite is not None:
            component = EstimateItem(
                section=current_section,
                number="",
                name=name_str,
                unit=str(unit) if unit else "",
                quantity=_to_float(qty) or 0,
                kind="material",
                vat_included=True,
                raw_row_index=idx + 1,
            )
            current_composite.components.append(component)
            continue

        if _is_section_header(row):
            current_section = name_str
            current_composite = None
            continue

        has_mat = _to_float(price_mat) is not None or _to_float(sum_mat) is not None
        has_work = _to_float(price_work) is not None or _to_float(sum_work) is not None

        is_composite_header = name_str.lower().endswith("в составе:")

        if is_composite_header:
            kind = "composite"
        elif has_mat and has_work:
            kind = "mixed"
        elif has_work and not has_mat:
            kind = "work"
        else:
            kind = "material"

        item = EstimateItem(
            section=current_section,
            number=str(num) if num is not None else "",
            name=name_str,
            unit=str(unit) if unit else "",
            quantity=_to_float(qty) or 0,
            kind=kind,
            price_material=_to_float(price_mat),
            price_work=_to_float(price_work),
            sum_material=_to_float(sum_mat),
            sum_work=_to_float(sum_work),
            sum_total=_to_float(sum_total),
            vat_included=True,
            raw_row_index=idx + 1,
        )
        estimate.items.append(item)
        current_composite = item if kind == "composite" else None

    return estimate
