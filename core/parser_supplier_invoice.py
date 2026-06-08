"""Парсер счёта поставщика через Claude API.

Принимает PDF или Excel, отправляет содержимое в Claude, получает structured JSON
со списком позиций. Использует tool_use чтобы гарантировать структуру.

Для PDF — используем нативное чтение PDF в Anthropic API (content type "document").
Для Excel — выгружаем содержимое в plain text и отправляем как user message.
"""
from __future__ import annotations

import base64
import os
from pathlib import Path
from typing import Optional

import openpyxl
from anthropic import Anthropic

from .models import SupplierInvoice, SupplierInvoiceItem


DEFAULT_MODEL = "claude-sonnet-4-5"


SYSTEM_PROMPT = """\
Ты — парсер счетов от поставщиков строительно-монтажных материалов и оборудования. \
Тебе дают счёт (PDF или текст из Excel). Извлеки из него ВСЕ строки-позиции \
(не итоги, не шапку, не подписи) и общую информацию.

Что важно:
1. Имя поставщика и его ИНН — из шапки счёта.
2. Номер счёта и его дата.
3. «Тег проекта» (project_tag) — ТОЛЬКО если в счёте есть явное поле \
«Примечание», «Назначение», «Объект», «Проект» или аналогичное с коротким \
произвольным текстом, идентифицирующим объект/проект. У большинства поставщиков \
такого поля НЕТ — в этом случае верни null. НЕ выдумывай тег, НЕ бери его из \
адреса, ИНН, названия товара, графы «Назначение платежа» или из юридических \
реквизитов. Если сомневаешься — верни null.
4. Для каждой позиции: line_no (номер п/п в счёте), name (наименование товара), \
article_supplier (артикул/код поставщика), article_manufacturer (артикул производителя — \
часто в отдельном столбце), unit (ед. изм.), quantity (кол-во), unit_price \
(цена за ед., как указано в счёте), vat_rate (0.22 или 0.20 или null).
5. vat_included: True если цена в счёте УЖЕ с НДС, False если без НДС. \
Обычно в счетах поставщиков пишут «Цена руб.» без НДС, отдельно «НДС», отдельно \
«Всего с НДС». Тогда vat_included = False.
6. Итоги: total_without_vat (итого без НДС) и total_with_vat (итого с НДС).

Возвращай ответ ТОЛЬКО через инструмент `submit_invoice`. Не пиши свободный текст.
"""


INVOICE_TOOL = {
    "name": "submit_invoice",
    "description": "Передать распарсенный счёт поставщика.",
    "input_schema": {
        "type": "object",
        "properties": {
            "supplier_name": {"type": "string"},
            "supplier_inn": {"type": ["string", "null"]},
            "invoice_number": {"type": ["string", "null"]},
            "invoice_date": {
                "type": ["string", "null"],
                "description": "Дата счёта в формате YYYY-MM-DD",
            },
            "project_tag": {
                "type": ["string", "null"],
                "description": "Тег/примечание проекта, если указано",
            },
            "vat_included_default": {"type": "boolean"},
            "total_without_vat": {"type": ["number", "null"]},
            "total_with_vat": {"type": ["number", "null"]},
            "items": {
                "type": "array",
                "items": {
                    "type": "object",
                    "properties": {
                        "line_no": {"type": ["integer", "null"]},
                        "name": {"type": "string"},
                        "article_supplier": {"type": ["string", "null"]},
                        "article_manufacturer": {"type": ["string", "null"]},
                        "unit": {"type": ["string", "null"]},
                        "quantity": {"type": "number"},
                        "unit_price": {"type": "number"},
                        "vat_rate": {"type": ["number", "null"]},
                        "vat_included": {"type": "boolean"},
                    },
                    "required": ["name", "quantity", "unit_price", "vat_included"],
                },
            },
        },
        "required": ["supplier_name", "items"],
    },
}


def _excel_to_text(file_path: Path) -> str:
    """Выгружаем все непустые ячейки из Excel в plain text для отправки в Claude."""
    wb = openpyxl.load_workbook(file_path, data_only=True)
    chunks = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        chunks.append(f"=== Лист: {sheet} ===")
        for row in ws.iter_rows(values_only=True):
            row_str = " | ".join(str(c) if c is not None else "" for c in row)
            if row_str.strip(" |"):
                chunks.append(row_str)
    return "\n".join(chunks)


def parse_invoice(file_path: str | Path, *, model: Optional[str] = None
                   ) -> SupplierInvoice:
    """Главная функция — на вход путь к PDF или XLSX, на выход — SupplierInvoice."""
    path = Path(file_path)
    model = model or os.getenv("CLAUDE_MODEL") or DEFAULT_MODEL
    client = Anthropic()

    suffix = path.suffix.lower()
    if suffix == ".pdf":
        pdf_data = base64.standard_b64encode(path.read_bytes()).decode("utf-8")
        user_content = [
            {
                "type": "document",
                "source": {
                    "type": "base64",
                    "media_type": "application/pdf",
                    "data": pdf_data,
                },
            },
            {
                "type": "text",
                "text": "Распарси этот счёт и верни результат через инструмент.",
            },
        ]
    elif suffix in (".xlsx", ".xls"):
        text = _excel_to_text(path)
        user_content = [{
            "type": "text",
            "text": (
                "Это текстовая выгрузка счёта поставщика из Excel. "
                "Распарси и верни результат через инструмент.\n\n"
                f"{text}"
            ),
        }]
    else:
        raise ValueError(f"Неподдерживаемый формат файла: {suffix}")

    response = client.messages.create(
        model=model,
        max_tokens=8000,
        system=SYSTEM_PROMPT,
        tools=[INVOICE_TOOL],
        tool_choice={"type": "tool", "name": "submit_invoice"},
        messages=[{"role": "user", "content": user_content}],
    )

    tool_use = next((b for b in response.content if b.type == "tool_use"), None)
    if tool_use is None:
        raise RuntimeError(
            f"Модель не вызвала submit_invoice. Ответ: {response.content!r}"
        )

    data = tool_use.input

    items = [
        SupplierInvoiceItem(
            line_no=it.get("line_no"),
            name=it["name"],
            article_supplier=it.get("article_supplier"),
            article_manufacturer=it.get("article_manufacturer"),
            unit=it.get("unit"),
            quantity=float(it["quantity"]),
            unit_price=float(it["unit_price"]),
            vat_rate=it.get("vat_rate"),
            vat_included=bool(it.get("vat_included", False)),
        )
        for it in data.get("items", [])
    ]

    return SupplierInvoice(
        supplier_name=data["supplier_name"],
        supplier_inn=data.get("supplier_inn"),
        invoice_number=data.get("invoice_number"),
        invoice_date=data.get("invoice_date"),
        project_tag=data.get("project_tag"),
        vat_included_default=bool(data.get("vat_included_default", False)),
        items=items,
        total_without_vat=data.get("total_without_vat"),
        total_with_vat=data.get("total_with_vat"),
        source_file=path.name,
    )
